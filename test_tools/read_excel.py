from openpyxl import load_workbook
from test_tools.read_config import ReadConfig   # 导入读取配置文件的类
from test_tools.gain_path import *      # 导入配置路径的文件
import pandas as pd         # 导入parser处理excel数据
from test_tools.reflect import Method


class DoExcel:  # 处理excel的类

    @classmethod  # 当在方法之中调用另一个方法时，使用类方法装饰
    def get_excel(cls, file_name):   # 读取excel文件的方法
        wb = load_workbook(file_name)   # 读取文件
        # 读取配置文件，文件数据是字典，读取出来是字符串此处需要转换，传入配置文件路径，区域名,选项值
        mode = eval(ReadConfig.get_config(test_config_file, "MODE", "mode"))
        amount = getattr(Method, 'amount')
        excel_data = []  # 创建列表嵌套字典的数据形式
        for key in mode:  # 遍历配置文件数据
            sheet = wb[key]  # 读取单元
            if mode[key] == 'all':  # 如果表单值等于all 就把表单数据全部加载到列表嵌套字典中
                for i in range(2, sheet.max_row + 1):
                    data = { }
                    data['case_id'] = sheet.cell(i, 1).value
                    data['module'] = sheet.cell(i, 2).value
                    data['title'] = sheet.cell(i, 3).value
                    data['url'] = sheet.cell(i, 4).value
                    if sheet.cell(i, 5).value.find("${amount}") != -1:  # 从该行查找子字符串，-1是找到就替换
                        data['data'] = sheet.cell(i, 5).value.replace("${amount}", str(amount))
                        # setattr(Method, 'amount', getattr(Method, 'amount') + 1) 适合于注册，每个手机号+1，此处暂时不用
                    else:
                        data['data'] = sheet.cell(i, 5).value
                    data['headers'] = sheet.cell(i, 6).value
                    data['method'] = sheet.cell(i, 7).value
                    data['expected_result'] = sheet.cell(i, 8).value
                    data['check_sql'] = sheet.cell(i, 9).value
                    data['result'] = sheet.cell(i, 10).value
                    data['sheet_name'] = key
                    excel_data.append(data)
                    cls.updata_price(amount, file_name, 'init')
                    # 初始化数据每次增加数量：用例的最大行-2，一行标题一行重复数据

            else:  # 否则的话，遍历key的值，如果表单case_id与列表相等，就把它加入到列表嵌套字典当中
                for case_id in mode[key]:
                    data = {}
                    data['case_id'] = sheet.cell(case_id+1, 1).value
                    data['module'] = sheet.cell(case_id+1, 2).value
                    data['title'] = sheet.cell(case_id+1, 3).value
                    data['url'] = sheet.cell(case_id+1, 4).value
                    if sheet.cell(case_id+1, 5).value.find("${amount}") != -1:  # 从该行查找子字符串，有找到
                        data['data'] = sheet.cell(case_id+1, 5).value.replace("${amount}", str(amount))
                        # setattr(Method, 'amount', getattr(Method, 'amount') + 1) 适合于注册，每个手机号+1，此处暂时不用
                    else:
                        data['data'] = sheet.cell(case_id+1, 5).value
                        # 如果是注册，此处可以添加一个写会的范围，根据总行数减去不成功用例
                        # 例如：sheet.max_row - 2
                    cls.updata_price(amount, file_name, 'init')
                    data['headers'] = sheet.cell(case_id+1, 6).value
                    data['method'] = sheet.cell(case_id+1, 7).value
                    data['expected_result'] = sheet.cell(case_id+1, 8).value
                    if sheet.cell(case_id+1, 9).value !=None:
                        data['check_sql'] = sheet.cell(case_id+1, 9).value
                    data['result'] = sheet.cell(case_id+1, 10).value
                    data['sheet_name'] = key
                    excel_data.append(data)

        return excel_data   # 返回列表数据

    @classmethod  # 把最新测试数据初始化数据写回excel，该方法在上个方法中调用，所以把它们都标记为类方法
    def updata_price(cls, price, filename, sheetname):  # 传入excel文件名称，表单名称，价格
        wb = load_workbook(filename)
        sheet = wb[sheetname]
        sheet.cell(2, 1).value = price  # 传入的price 赋值给把指定位置的第2行第1列
        wb.save(filename)  # 保存文件

    @staticmethod  # 把测试数据写回excel的方法
    def write_back(file_name, sheet_name, i, j, result):  # 传入文件，表单，行，测试结果，对比结果
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, j).value = result  # 把测试结果写到第10列
        wb.save(file_name)          # 保存结果

    @staticmethod  # 读取特定列的值，因为需要我们只需要读一条订单号，此处把列写死了，
    def read_value(file_name, sheet_name, u):       # 传入文件名称，表单，行
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        bos = eval(sheet.cell(u, 10).value)  # 返回的是字典，取出来变成字符串，转换一下
        ws = bos['data']['order_no']
        return ws           # 返回我们需要的订单号


if __name__ == '__main__':
    DoExcel.get_excel(test_excel_path)


